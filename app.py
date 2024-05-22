from flask import Flask, render_template, request
import pandas as pd
import re
import openpyxl
from openpyxl import load_workbook
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Blueprint



app = Flask(__name__)


def check_password(password, repassword):
    if password == None and repassword == None:
        return False
    elif password != repassword:
        return False
    elif len(password) < 7:
        return False
    elif not re.search(r'\d', password):
        return False
    elif not re.search(r'[A-Z]', password):
        return False
    else:
        return True


def check_name(name):
    if 0<len(name) > 31:
        return False
    elif name==None:
        return False
    elif not re.match(r'^[a-zA-Zа-яА-Я]+$', name):
        return False
    else:
        return True


def check_email(email):
    if not re.search(r'@', email):
        return False
    elif email == None:
        return False
    else:
        return True


def submit():
    data = {
        'name': [request.form['name']],
        'email': [request.form['email']],
        'password': [request.form['password']],
    }
    df = pd.DataFrame(data)
    df.to_excel('users.xlsx', index=False)
    return 'Успешная регистрация!'


@app.route('/check_reg', methods=['POST'])
def account_creation():
    name = request.form['name']
    email = request.form['email']
    password = request.form['password']
    repassword = request.form['repassword']
    if(check_name(name) and check_password(password, repassword) and check_email(email)):
        return submit()
    else:
        return 'Провальная попытка..'


def check_login(name, email, password):
    # Открываем файл Excel
    wb = load_workbook('users.xlsx')
    sheet = wb.active

    # Проверяем данные
    for row in sheet.iter_rows(min_row=2, values_only=True):
        saved_name, saved_email, saved_password = row
        if name == saved_name and email == saved_email and password == saved_password:
            return True
        else:
            return False


@app.route('/check_auth', methods=['POST'])
def login():
    name = request.form['name']
    email = request.form['email']
    password = request.form['password']

    if check_login(name, email, password):
        return 'Авторизация успешна!'
    else:
        return 'Неверные учетные данные..'


@app.route('/')
def registration():
    return render_template('registration.html')


@app.route('/auth')
def authorisation():
    return render_template("authorisation.html")


if __name__ == '__main__':
    app.run(debug=False)